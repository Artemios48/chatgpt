import telebot
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from telebot import types

fn = 'db.xlsx'
wb = load_workbook(fn)
ws = wb['data']
print(f'Бот запущен a{ws.max_row}')

bot = telebot.TeleBot("")
@bot.message_handler(content_types=["text"])

def start(message):
    if message.text == "/start":
        markup = types.ReplyKeyboardMarkup(resize_keyboard = True)
        btn_adm = types.KeyboardButton('Код администрации')
        btn_1 = types.KeyboardButton('Предложение\замечание')
        markup.add(btn_adm, btn_1)
        bot.send_message(message.from_user.id,'Привет! 👋  Я — «Комфортная гармония», ваш помощник в гимназии «Гармония»! Мы вместе с вами делаем школу комфортнее и уютнее, участвуя в проекте «Комфортная школа». Отправляйте мне ваши идеи, предложения и замечания — вместе мы сделаем «Гармонию» еще лучше! 😊', reply_markup=markup)
        bot.register_next_step_handler(message, change)
    else:
        bot.send_message(message.from_user.id,'Напишите /start для начала работы')
def change(message):
    id = str(message.from_user.id)
    if message.text == 'Предложение\замечание':
        for row in range(2,ws.max_row+1):
            id_ = ws.cell(row=row, column=7).value
            if id_ == id:
                name = ws.cell(row=row, column=1).value
                surname = ws.cell(row=row, column=2).value
                username = ws.cell(row=row, column=3).value
                print(name,surname,username,id)#удалить
                bot.send_message(message.from_user.id,'У вас появились ещё идеи? Пишите ваше предложение\замечание')
                bot.register_next_step_handler(message, lambda msg: get_inf_st(msg, name, surname, username))
                break
        else:
            bot.send_message(message.from_user.id,'Отправьте ваше сообщение по следующей форме: Имя Фамилия, предложение\замечание')
            bot.register_next_step_handler(message, get_inf)
    elif message.text == 'Код администрации':
        bot.send_message(message.from_user.id,'Отпарвьте ваш код администрации для просмотра запросов')
        bot.register_next_step_handler(message, get_adm)
    else:
        bot.send_message(message.from_user.id,'Извините, выберите вариант из предложенных на клавиатуре.')
        bot.register_next_step_handler(message, change)

def get_inf_st(message, name, surname, username):
    _1 = message.text
    time = datetime.now().strftime("%H:%M:%S %d.%m.%Y")
    id = str(message.from_user.id)
    ws.append([name,surname,username,_1,time,'Не просмотрено',id])
    wb.save(fn)
    wb.close()
    bot.send_message(message.chat.id, 'Спасибо за ваше сообщение! Ваше предложение/замечание успешно записано.')
    bot.register_next_step_handler(message, change)

def get_inf(message):
    id = str(message.from_user.id)
    username = str(message.from_user.username)
    time = datetime.now().strftime("%H:%M:%S %d.%m.%Y")
    try:
        name_surname, _1 = map(str,message.text.split(','))
        name,surname = name_surname.split(' ')
        ws.append([name,surname,username,_1,time,'Не просмотрено',id])
        wb.save(fn)
        wb.close()
        bot.send_message(message.chat.id, 'Спасибо за ваше сообщение! Ваше предложение/замечание успешно записано.')
        bot.register_next_step_handler(message, change)
    except:
        bot.send_message(message.chat.id, 'Вы отправили сообщение не по форме указанной выше! Попробуйте ещё раз.')
        bot.register_next_step_handler(message, get_inf)

def get_adm(message):
    if message.text == '12345':
        bot.send_message(message.from_user.id,'Режим администрации активен')
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn_send_db = types.KeyboardButton('Получить отчёт')
        btn_view = types.KeyboardButton('Просмотреть идеи')
        btn_exit = types.KeyboardButton('Выход')
        markup.add(btn_send_db, btn_view, btn_exit)
        bot.send_message(message.chat.id, 'Выберите действие', reply_markup=markup)
        bot.register_next_step_handler(message, handle_adm_choice)
    else:
        bot.send_message(message.from_user.id,'Код администрации введён не верно')
        bot.register_next_step_handler(message, get_adm)

def handle_adm_choice(message):
    if message.text == 'Просмотреть идеи':
        show_ideas(message)
        bot.register_next_step_handler(message, handle_adm_choice)
    elif message.text == 'Получить отчёт':
        send_db(message)
        bot.register_next_step_handler(message, handle_adm_choice)
    elif message.text == 'Выход':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn_adm = types.KeyboardButton('Код администрации')
        btn_1 = types.KeyboardButton('Предложение\замечание')
        markup.add(btn_adm, btn_1)
        bot.send_message(message.chat.id, 'Вы вышли из режима администрации.', reply_markup=markup) 
        bot.register_next_step_handler(message, change)
    else:
        bot.send_message(message.from_user.id,'Извините, выберите вариант из предложенных на клавиатуре.')
        bot.register_next_step_handler(message, handle_adm_choice)
    


def show_ideas(message):
    has_new_ideas = False
    for row in range(2,ws.max_row+1):
        status = ws.cell(row=row, column=6).value
        if status == 'Не просмотрено':
            has_new_ideas = True
            name = ws.cell(row=row, column=1).value
            surname = ws.cell(row=row, column=2).value
            username = ws.cell(row=row, column=3).value
            problem = ws.cell(row=row, column=4).value
            time = ws.cell(row=row, column=5).value
            id = ws.cell(row=row, column=7).value
        if status == 'Не просмотрено':
            if username == 'None':
                bot.send_message(message.chat.id,f'{name} {surname} \n https://web.telegram.org/k/#{id} \n ⁉️{problem} \n 🕒{time} \n 🔔{status}')
            else:
                bot.send_message(message.chat.id,f'{name} {surname} \n @{username} \n ⁉️{problem} \n 🕒{time} \n 🔔{status}')
            ws.cell(row=row, column=6).value = 'Просмотрено'  # Обновляем статус
            wb.save(fn)
    if not has_new_ideas:
        bot.send_message(message.chat.id, "Новых идей пока нет.")

def send_db(message):
    with open(fn, 'rb') as file:
        bot.send_document(message.chat.id, file)
        

bot.polling(none_stop=True, interval = 0)